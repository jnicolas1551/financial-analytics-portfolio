# =============================================================================
# DESCARGA DE PRECIOS DIARIOS - ACCIONES + S&P 500
# Fuente  : Yahoo Finance via yfinance
# Output  : Hoja "BASE DE DATOS LIMPIA" del Excel de análisis
# Moneda  : USD para todos los activos
# =============================================================================
#
# ACTIVOS INCLUIDOS:
#   ^GSPC  → S&P 500 (índice)
#   AAPL   → Apple
#   KO     → Coca-Cola
#   EC     → Ecopetrol (ADR en USD)
#   NVDA   → Nvidia
#   AMZN   → Amazon
#   TSLA   → TESLA
#
# LÓGICA DE LIMPIEZA:
#   1. Descarga precios de cierre ajustado (Adj Close) para cada ticker
#   2. Elimina filas donde CUALQUIER activo no tiene dato
#   3. Genera log auditable con detalle de filas eliminadas
#   4. Escribe datos limpios directamente en el Excel
#
# CÓMO CORRER:
#   1. pip install yfinance openpyxl pandas
#   2. Coloca este script en la misma carpeta que tu Excel
#   3. python descarga_acciones.py
#
# CÓMO MODIFICAR ACTIVOS:
#   Solo cambia la sección CONFIGURACIÓN → TICKERS
#   El resto del script se adapta solo
#
# =============================================================================


# -----------------------------------------------------------------------------
# SECCIÓN 1: LIBRERÍAS
# -----------------------------------------------------------------------------
import yfinance as yf        # Descarga de datos Yahoo Finance
import pandas as pd          # Manipulación de datos
import openpyxl              # Escritura en Excel preservando formato
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
import sys


# -----------------------------------------------------------------------------
# SECCIÓN 2: CONFIGURACIÓN — AQUÍ MODIFICAS SIN TOCAR EL RESTO
# -----------------------------------------------------------------------------

# Ruta completa del archivo Excel
ARCHIVO_EXCEL = r"D:\Datos Usuario\Documents\PROYECTO APRENDIZAJE IA\1 FASE CLAUDE\AUTO. DESCARGA ACCIONES\PYTHON_AUTOMATICO.xlsx"

# Hoja donde se pegan los datos
HOJA_DATOS = "BASE DE DATOS LIMPIA"

# Fecha de inicio de descarga
FECHA_INICIO = "2020-01-01"

# Tickers a descargar → formato: {"NOMBRE_COLUMNA": "TICKER_YAHOO"}
# Para cambiar activos: modifica este diccionario
TICKERS = {
    "S&P 500":   "^GSPC",
    "APPLE":     "AAPL",
    "COCA-COLA": "KO",
    "ECOPETROL": "EC",
    "NVIDIA":    "NVDA",
    "AMAZON":    "AMZN",
"TESLA":    "TSLA",	
}

# Nombre del archivo de log auditable
ARCHIVO_LOG = "log_descarga.txt"


# -----------------------------------------------------------------------------
# SECCIÓN 3: FUNCIONES
# -----------------------------------------------------------------------------

def log(mensaje, archivo_log, imprimir=True):
    """Escribe mensaje en consola y en archivo de log."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    linea = f"[{timestamp}] {mensaje}"
    if imprimir:
        print(linea)
    with open(archivo_log, "a", encoding="utf-8") as f:
        f.write(linea + "\n")


def descargar_datos(tickers, fecha_inicio, archivo_log):
    """
    Descarga precios de cierre ajustado de Yahoo Finance.
    Retorna DataFrame con fechas en índice y tickers como columnas.
    """
    print("\n" + "="*60)
    print("MÓDULO 1: DESCARGA DE DATOS")
    print("="*60)

    simbolos = list(tickers.values())
    nombres  = list(tickers.keys())

    log(f"Descargando {len(simbolos)} activos desde {fecha_inicio}", archivo_log)
    log(f"Activos: {simbolos}", archivo_log)

    # yfinance descarga todos en una sola llamada (eficiente)
    raw = yf.download(
        tickers    = simbolos,
        start      = fecha_inicio,
        auto_adjust= True,        # Precios ajustados por dividendos y splits
        progress   = True
    )

    # Extraer solo columna "Close" y renombrar con nombres legibles
    if len(simbolos) > 1:
        df = raw["Close"].copy()
        # Reordenar columnas según orden del diccionario
        df = df[simbolos]
        df.columns = nombres
    else:
        df = raw["Close"].to_frame()
        df.columns = nombres

    df.index.name = "fecha_corte"
    log(f"Descarga completa: {len(df)} filas brutas", archivo_log)
    log(f"Rango bruto: {df.index.min().date()} → {df.index.max().date()}", archivo_log)

    return df


def limpiar_datos(df, archivo_log):
    """
    Elimina filas con cualquier valor faltante.
    Genera log detallado de qué se eliminó y por qué.
    Retorna DataFrame limpio.
    """
    print("\n" + "="*60)
    print("MÓDULO 2: LIMPIEZA Y AUDITORÍA")
    print("="*60)

    filas_brutas = len(df)
    log(f"Filas antes de limpieza: {filas_brutas}", archivo_log)

    # Identificar filas con NaN y documentar cuáles y por qué
    filas_con_nan = df[df.isnull().any(axis=1)]

    if len(filas_con_nan) > 0:
        log(f"Filas eliminadas: {len(filas_con_nan)}", archivo_log)
        log("Detalle de filas eliminadas:", archivo_log)

        for fecha, fila in filas_con_nan.iterrows():
            activos_sin_dato = fila[fila.isnull()].index.tolist()
            log(f"  {fecha.date()} → sin dato en: {activos_sin_dato}", archivo_log)
    else:
        log("Sin filas eliminadas: todos los activos tienen datos completos", archivo_log)

    # Limpieza: eliminar cualquier fila con NaN
    df_limpio = df.dropna()

    filas_limpias = len(df_limpio)
    log(f"Filas después de limpieza: {filas_limpias}", archivo_log)
    log(f"Rango final: {df_limpio.index.min().date()} → {df_limpio.index.max().date()}", archivo_log)

    # Resumen de completitud por activo
    print("\nCOMPLETITUD POR ACTIVO (datos limpios):")
    for col in df_limpio.columns:
        print(f"  {col:<15}: {df_limpio[col].notna().sum():>5} fechas")

    return df_limpio


def escribir_excel(df, archivo_excel, hoja_datos, archivo_log):
    """
    Escribe el DataFrame limpio en la hoja BASE DE DATOS LIMPIA del Excel.
    Preserva el resto de hojas y su formato intacto.
    """
    print("\n" + "="*60)
    print("MÓDULO 3: ESCRITURA EN EXCEL")
    print("="*60)

    if not os.path.exists(archivo_excel):
        log(f"ERROR: No se encontró el archivo {archivo_excel}", archivo_log)
        log("Verifica que el Excel esté en la misma carpeta que este script", archivo_log)
        sys.exit(1)

    log(f"Abriendo: {archivo_excel}", archivo_log)
    wb = openpyxl.load_workbook(archivo_excel)

    if hoja_datos not in wb.sheetnames:
        log(f"ERROR: No existe la hoja '{hoja_datos}' en el archivo", archivo_log)
        log(f"Hojas disponibles: {wb.sheetnames}", archivo_log)
        sys.exit(1)

    ws = wb[hoja_datos]

    # Limpiar contenido existente (sin borrar el formato)
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    log("Hoja limpiada. Escribiendo nuevos datos...", archivo_log)

    # --- Estilos ---
    fuente_header = Font(bold=True, color="FFFFFF", size=10)
    relleno_header = PatternFill("solid", start_color="1F4E79")  # Azul oscuro
    alineacion_centro = Alignment(horizontal="center", vertical="center")
    borde_fino = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    relleno_alterno = PatternFill("solid", start_color="EBF3FB")  # Azul muy claro

    # --- Fila 1: Headers ---
    headers = ["fecha_corte"] + list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = fuente_header
        cell.fill = relleno_header
        cell.alignment = alineacion_centro
        cell.border = borde_fino

    # Ajustar ancho de columnas
    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, len(headers) + 1):
        col_letra = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letra].width = 14

    # --- Filas de datos ---
    for fila_idx, (fecha, fila) in enumerate(df.iterrows(), start=2):
        # Fecha
        cell_fecha = ws.cell(row=fila_idx, column=1, value=fecha.date())
        cell_fecha.number_format = "YYYY-MM-DD"
        cell_fecha.alignment = alineacion_centro
        cell_fecha.border = borde_fino

        # Alternar color de filas para legibilidad
        if fila_idx % 2 == 0:
            cell_fecha.fill = relleno_alterno

        # Precios
        for col_idx, valor in enumerate(fila, start=2):
            cell = ws.cell(row=fila_idx, column=col_idx, value=round(float(valor), 6))
            cell.number_format = "#,##0.000000"
            cell.alignment = alineacion_centro
            cell.border = borde_fino
            if fila_idx % 2 == 0:
                cell.fill = relleno_alterno

    # Fijar fila de headers al hacer scroll
    ws.freeze_panes = "A2"

    # --- Celda de auditoría ---
    # Se escribe en la primera celda disponible debajo de los datos
    fila_auditoria = len(df) + 3
    ws.cell(row=fila_auditoria, column=1,
            value=f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')} | "
                  f"Filas: {len(df)} | "
                  f"Rango: {df.index.min().date()} → {df.index.max().date()} | "
                  f"Fuente: Yahoo Finance (USD, precios ajustados)"
    ).font = Font(italic=True, color="888888", size=8)

    wb.save(archivo_excel)
    log(f"Excel guardado: {os.path.abspath(archivo_excel)}", archivo_log)
    log(f"Filas escritas: {len(df)}", archivo_log)


# -----------------------------------------------------------------------------
# SECCIÓN 4: EJECUCIÓN PRINCIPAL
# -----------------------------------------------------------------------------

if __name__ == "__main__":

    # Inicializar log (sobrescribe el anterior para mantener solo el último)
    with open(ARCHIVO_LOG, "w", encoding="utf-8") as f:
        f.write(f"{'='*60}\n")
        f.write(f"LOG DE DESCARGA - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"{'='*60}\n\n")

    print("="*60)
    print("PIPELINE: YAHOO FINANCE → EXCEL")
    print("="*60)

    # PASO 1: Descargar
    df_bruto = descargar_datos(TICKERS, FECHA_INICIO, ARCHIVO_LOG)

    # PASO 2: Limpiar y auditar
    df_limpio = limpiar_datos(df_bruto, ARCHIVO_LOG)

    # PASO 3: Escribir en Excel
    escribir_excel(df_limpio, ARCHIVO_EXCEL, HOJA_DATOS, ARCHIVO_LOG)

    print("\n" + "="*60)
    print("PROCESO COMPLETADO")
    print(f"  Excel actualizado : {ARCHIVO_EXCEL}")
    print(f"  Log auditable     : {ARCHIVO_LOG}")
    print("="*60)


# =============================================================================
# NOTAS PARA MODIFICAR EL SCRIPT
# =============================================================================
#
# AGREGAR UN ACTIVO NUEVO:
#   En TICKERS, agrega una línea:
#   "NOMBRE_COLUMNA": "TICKER_YAHOO"
#   Ejemplo: "TESLA": "TSLA"
#   El script lo incluye automáticamente en todo el proceso.
#
# CAMBIAR FECHA DE INICIO:
#   Modifica FECHA_INICIO = "2016-01-01" por la fecha que necesites.
#
# CAMBIAR EL ARCHIVO EXCEL:
#   Modifica ARCHIVO_EXCEL con el nombre exacto de tu archivo.
#   El archivo debe estar en la misma carpeta que el script.
#
# ENTENDER EL LOG:
#   Cada vez que corres el script se genera log_descarga.txt
#   Ahí encuentras exactamente qué fechas se eliminaron y por qué.
#   Ese archivo es tu evidencia auditable del proceso.
#
# REQUISITOS:
#   pip install yfinance openpyxl pandas
#
# =============================================================================
